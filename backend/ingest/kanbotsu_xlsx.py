"""国土交通省 道路陥没オープンデータ（令和5・6年度、約22,000件）の取込。

地名文字列（「〇〇市〇〇地先」等）を国土地理院 地名検索APIでジオコーディングし、
SubsidenceEventとして保存する。位置精度は字・大字レベルまでで、番地は解決しない
――これはMonitorの「検証フック」（陥没とInSAR沈下/地盤の空間相関）に使う教師データ
であり、個別管路の位置特定に使うものではない。

シート構成・列位置は年度によって微妙に異なる（例: R6の都道府県シートには
R5に無い列が1つ追加されている）ため、列位置は固定せずヘッダー行のラベルから
動的に検出する。このデータセットにはイベントごとの正確な発生日は含まれない
（年度単位のみ）。
"""
from __future__ import annotations

import time
from io import BytesIO

import requests
from django.contrib.gis.geos import Point
from openpyxl import load_workbook

from core.mesh import mesh_cell_for_geometry
from core.models import SubsidenceEvent

REQUEST_HEADERS = {"User-Agent": "SinkScope/0.1 (+https://github.com/ops324/sinkscope)"}
GEOCODE_URL = "https://msearch.gsi.go.jp/address-search/AddressSearch"

SOURCE_FILES = {
    "r5": "https://www.mlit.go.jp/road/sisaku/ijikanri/xlsx/dourokanbotsu/r5/02.xlsx",
    "r6": "https://www.mlit.go.jp/road/sisaku/ijikanri/xlsx/dourokanbotsu/r6/02.xlsx",
}

TARGET_PREFECTURES = ("東京都", "埼玉県")

# ヘッダーはタイトル行・空行を挟んだ2行構成(主要項目 / 「大きさ」の内訳)で、
# 実データはそれらの後、常に(0-indexで)6行目から始まる(r5/r6・3ジュリスディクション
# 全パターンで確認済み)。
_HEADER_SCAN_ROWS = 6
_DATA_START_INDEX = 6
_REQUIRED_LABELS = ("道路管理者", "路線名", "地名", "要因施設", "深さ(m)")
_OPTIONAL_LABELS = ("縦断方向(m)", "横断方向(m)")


def _jurisdiction_for_sheet(sheet_name: str) -> str | None:
    if "直轄" in sheet_name:
        return "national"
    if "都道府県" in sheet_name:
        return "prefectural"
    if "市町村" in sheet_name:
        return "municipal"
    return None


def _find_column_positions(rows: list[tuple]) -> dict[str, int]:
    """ヘッダー行の中からラベルを探して列位置を特定する(固定位置に依存しない)。"""
    positions: dict[str, int] = {}
    for row in rows[:_HEADER_SCAN_ROWS]:
        for label in (*_REQUIRED_LABELS, *_OPTIONAL_LABELS):
            if label not in positions and label in row:
                positions[label] = row.index(label)

    missing = set(_REQUIRED_LABELS) - positions.keys()
    if missing:
        raise ValueError(f"想定した列見出しが見つかりません: {missing}")
    return positions


def _iter_data_rows(rows: list[tuple], positions: dict[str, int]):
    for row in rows[_DATA_START_INDEX:]:
        location_text = row[positions["地名"]]
        if not location_text:
            continue
        yield {
            "road_administrator": str(row[positions["道路管理者"]] or "").strip(),
            "road_name": str(row[positions["路線名"]] or "").strip(),
            "location_text": str(location_text).strip(),
            "cause_facility": str(row[positions["要因施設"]] or "").strip(),
            "depth_m": row[positions["深さ(m)"]],
        }


def _download_workbook(url: str):
    response = requests.get(url, headers=REQUEST_HEADERS, timeout=60)
    response.raise_for_status()
    return load_workbook(BytesIO(response.content), data_only=True)


def _geocode(location_text: str) -> tuple[Point | None, str]:
    """地名文字列を国土地理院 地名検索APIでジオコーディングする(字・大字レベル)。"""
    try:
        response = requests.get(
            GEOCODE_URL, params={"q": location_text}, headers=REQUEST_HEADERS, timeout=15
        )
        response.raise_for_status()
        features = response.json()
    except (requests.RequestException, ValueError):
        return None, "failed"

    if not features:
        return None, "not_found"

    lon, lat = features[0]["geometry"]["coordinates"]
    return Point(lon, lat, srid=4326), "approximate"


def _to_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def ingest_subsidence_events(
    fiscal_years: tuple[str, ...] = ("r5", "r6"), geocode_delay_s: float = 1.0
) -> int:
    """国交省の陥没オープンデータを取込み、地名をジオコーディングして保存する。

    地名検索APIへの配慮として、明文化されたレート制限はないがリクエスト間隔を空ける。
    再実行しても重複しないよう、対象年度分を一度削除してから作り直す。
    """
    SubsidenceEvent.objects.filter(source__in=[f"mlit_kanbotsu_{fy}" for fy in fiscal_years]).delete()

    events = []
    for fiscal_year in fiscal_years:
        workbook = _download_workbook(SOURCE_FILES[fiscal_year])
        for sheet_name in workbook.sheetnames:
            jurisdiction = _jurisdiction_for_sheet(sheet_name)
            if jurisdiction is None:
                continue

            rows = list(workbook[sheet_name].iter_rows(values_only=True))
            positions = _find_column_positions(rows)

            for record in _iter_data_rows(rows, positions):
                if not record["location_text"].startswith(TARGET_PREFECTURES):
                    continue

                point, confidence = _geocode(record["location_text"])
                time.sleep(geocode_delay_s)
                mesh_cell = mesh_cell_for_geometry(point) if point else None

                events.append(
                    SubsidenceEvent(
                        geom=point,
                        mesh_cell=mesh_cell,
                        fiscal_year=fiscal_year,
                        prefecture=record["location_text"][:3],
                        location_text=record["location_text"],
                        road_name=record["road_name"],
                        road_administrator=record["road_administrator"],
                        road_jurisdiction=jurisdiction,
                        cause_facility=record["cause_facility"],
                        depth_m=_to_float(record["depth_m"]),
                        geocode_confidence=confidence,
                        source=f"mlit_kanbotsu_{fiscal_year}",
                    )
                )

    SubsidenceEvent.objects.bulk_create(events)
    return len(events)
